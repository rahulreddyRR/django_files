from openpyxl import load_workbook,workbook
wb = load_workbook('base.xlsx')
sheet = wb['Sheet1']

wx = workbook('base.xlse')
print(type(wx))

c = sheet.cell(row=1,column=1)
print(c.value)

c1 = sheet['B4'].value
print(sheet.max_column)

for row in range(1,sheet.max_row+1):
    for col in range(1,sheet.max_column+1):
        print(sheet.cell(row,column=col).value,end="   ")
    print()
